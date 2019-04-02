from PIL import Image
import openpyxl
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell


# im = Image.open('fnc.jpg')
# width, height = im.size
# newWidth = int(width/20)
# newHeight = int(height/20)
# wb = load_workbook('im.xlsx')
# ws = wb.active
# newIm = im.resize((newWidth, newHeight), Image.ANTIALIAS)
# pixel = newIm.getpixel((1, 1))
# str = '%02X%02X%02X' % (pixel[0], pixel[1], pixel[2])
# print(pixel)
# ws[1][1].fill = PatternFill(bgColor=str, fill_type='solid')
# wb.save('im.xlsx')
# # newIm.show
# # newIm.save('new.jpg', 'JPEG')
# # newIm.close()
# ##im.show()
#
# print(width)
# print(height)

def num2Col(number):
    letters = ''
    while number:
        mod = (number - 1) % 26
        letters += chr(mod + 65)
        number = (number - 1) // 26
    return ''.join(reversed(letters))
im = Image.open('fncc.jpg')
width, height = im.size

newWidth = int(width/10)
newHeight = int(height/10)
newIm = im.resize((newWidth, newHeight), Image.ANTIALIAS)
newIm.show

wb = load_workbook('im.xlsx')
ws = wb.active
for i in range (0,newWidth):
    for j in range (0,newHeight):
        pixel = newIm.getpixel((i, j))
        hex = '%02X%02X%02X' % (pixel[0], pixel[1], pixel[2])
        if j == 0:
            ws.column_dimensions[num2Col(i+1)].width = 4
        cellLocation = num2Col(i+1)+str(j+1)
        print(cellLocation + ' '+ hex)
        ws[cellLocation].fill = PatternFill(fgColor= hex, fill_type='solid')
wb.save('im.xlsx')